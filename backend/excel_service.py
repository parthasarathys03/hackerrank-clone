"""
Excel Service for Assessment Results
Handles reading, writing, and exporting Excel data
"""

import os
from datetime import datetime
from typing import Dict, List, Optional, Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Excel file path
EXCEL_FILE_PATH = os.path.join(os.path.dirname(__file__), "data", "assessment_results.xlsx")

# Sheet names
SHEET_TEST_SUMMARY = "test_summary"
SHEET_PROBLEM_TESTCASES = "problem_testcases"
SHEET_TESTCASE_DETAILS = "testcase_details"

# Column headers for each sheet
TEST_SUMMARY_COLUMNS = [
    "candidate_id", "name", "email", "phone", "test_date", "login_time", "submit_time",
    "submission_type", "time_taken_min", "total_questions", "python_questions",
    "sql_questions", "python_score", "sql_score", "overall_score",
    "python_score_percentage", "sql_score_percentage", "overall_percentage",
    "overall_verdict"
]

PROBLEM_TESTCASES_COLUMNS = [
    "candidate_id", "name", "easy_solved", "easy_total", "medium_solved", "medium_total",
    "hard_solved", "hard_total", "total_solved", "total_problems"
]

TESTCASE_DETAILS_COLUMNS = [
    "candidate_id", "name",
    "P1_py", "P2_py", "P3_py", "P4_py", "P5_py",
    "P6_sql", "P7_sql", "P8_sql", "P9_sql", "P10_sql"
]

# Colors for verdict
VERDICT_COLORS = {
    "Good": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # Light green
    "Average": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # Light yellow
    "Below Average": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
}


def calculate_verdict(overall_percentage: float) -> str:
    """Calculate verdict based on overall percentage"""
    if overall_percentage < 40:
        return "Below Average"
    elif overall_percentage <= 60:
        return "Average"
    else:
        return "Good"


def calculate_scores(python_score: float, sql_score: float) -> Dict[str, float]:
    """Calculate all derived scores and percentages"""
    overall_score = python_score + sql_score
    overall_percentage = (overall_score / 200) * 100
    python_score_percentage = (python_score / 100) * 100
    sql_score_percentage = (sql_score / 100) * 100
    overall_verdict = calculate_verdict(overall_percentage)
    
    return {
        "overall_score": overall_score,
        "overall_percentage": round(overall_percentage, 2),
        "python_score_percentage": round(python_score_percentage, 2),
        "sql_score_percentage": round(sql_score_percentage, 2),
        "overall_verdict": overall_verdict
    }


def ensure_excel_exists():
    """Create Excel file with headers if it doesn't exist"""
    if os.path.exists(EXCEL_FILE_PATH):
        return
    
    wb = Workbook()
    
    # Sheet 1: Test Summary
    ws1 = wb.active
    ws1.title = SHEET_TEST_SUMMARY
    for col, header in enumerate(TEST_SUMMARY_COLUMNS, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Sheet 2: Problem Testcases
    ws2 = wb.create_sheet(SHEET_PROBLEM_TESTCASES)
    for col, header in enumerate(PROBLEM_TESTCASES_COLUMNS, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Sheet 3: Testcase Details
    ws3 = wb.create_sheet(SHEET_TESTCASE_DETAILS)
    for col, header in enumerate(TESTCASE_DETAILS_COLUMNS, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Freeze top row on all sheets
    ws1.freeze_panes = "A2"
    ws2.freeze_panes = "A2"
    ws3.freeze_panes = "A2"
    
    wb.save(EXCEL_FILE_PATH)
    wb.close()


def read_all_results(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    verdict: Optional[str] = None,
    submission_type: Optional[str] = None
) -> List[Dict[str, Any]]:
    """Read all results from Excel with optional filters"""
    ensure_excel_exists()
    
    wb = load_workbook(EXCEL_FILE_PATH, read_only=True)
    
    # Read Sheet 1: Test Summary
    ws1 = wb[SHEET_TEST_SUMMARY]
    test_summary_data = {}
    headers1 = [cell.value for cell in ws1[1]]
    
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if row[0] is None:  # Skip empty rows
            continue
        record = dict(zip(headers1, row))
        candidate_id = record["candidate_id"]
        test_summary_data[candidate_id] = record
    
    # Read Sheet 2: Problem Testcases
    ws2 = wb[SHEET_PROBLEM_TESTCASES]
    problem_testcases_data = {}
    headers2 = [cell.value for cell in ws2[1]]
    
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        record = dict(zip(headers2, row))
        candidate_id = record["candidate_id"]
        problem_testcases_data[candidate_id] = {
            "easy_solved": record.get("easy_solved", 0),
            "easy_total": record.get("easy_total", 2),
            "medium_solved": record.get("medium_solved", 0),
            "medium_total": record.get("medium_total", 9),
            "hard_solved": record.get("hard_solved", 0),
            "hard_total": record.get("hard_total", 9),
            "total_solved": record.get("total_solved", 0),
            "total_problems": record.get("total_problems", 20)
        }
    
    # Read Sheet 3: Testcase Details
    ws3 = wb[SHEET_TESTCASE_DETAILS]
    testcase_details_data = {}
    headers3 = [cell.value for cell in ws3[1]]
    
    for row in ws3.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        record = dict(zip(headers3, row))
        candidate_id = record["candidate_id"]
        testcase_details_data[candidate_id] = {
            "P1_py": record.get("P1_py", 0),
            "P2_py": record.get("P2_py", 0),
            "P3_py": record.get("P3_py", 0),
            "P4_py": record.get("P4_py", 0),
            "P5_py": record.get("P5_py", 0),
            "P6_sql": record.get("P6_sql", 0),
            "P7_sql": record.get("P7_sql", 0),
            "P8_sql": record.get("P8_sql", 0),
            "P9_sql": record.get("P9_sql", 0),
            "P10_sql": record.get("P10_sql", 0)
        }
    
    wb.close()
    
    # Combine all data
    results = []
    for candidate_id, summary in test_summary_data.items():
        combined = {**summary}
        combined["problem_testcases"] = problem_testcases_data.get(candidate_id, {})
        combined["problem_scores"] = testcase_details_data.get(candidate_id, {})
        results.append(combined)
    
    # Apply filters
    filtered_results = []
    for record in results:
        # Date filter
        if date_from or date_to:
            test_date = record.get("test_date")
            if test_date:
                if isinstance(test_date, str):
                    test_date_obj = datetime.strptime(test_date, "%Y-%m-%d").date()
                else:
                    test_date_obj = test_date
                
                if date_from:
                    from_date = datetime.strptime(date_from, "%Y-%m-%d").date()
                    if test_date_obj < from_date:
                        continue
                
                if date_to:
                    to_date = datetime.strptime(date_to, "%Y-%m-%d").date()
                    if test_date_obj > to_date:
                        continue
        
        # Verdict filter
        if verdict and verdict != "All":
            if record.get("overall_verdict") != verdict:
                continue
        
        # Submission type filter
        if submission_type and submission_type != "All":
            if record.get("submission_type") != submission_type:
                continue
        
        filtered_results.append(record)
    
    return filtered_results


def add_result(data: Dict[str, Any]) -> Dict[str, Any]:
    """Add a new result to the Excel file"""
    ensure_excel_exists()
    
    # Calculate derived values
    python_score = data.get("python_score", 0)
    sql_score = data.get("sql_score", 0)
    calculated = calculate_scores(python_score, sql_score)
    
    # Merge calculated values
    data.update(calculated)
    
    wb = load_workbook(EXCEL_FILE_PATH)
    
    # Add to Sheet 1: Test Summary
    ws1 = wb[SHEET_TEST_SUMMARY]
    row1 = []
    for col in TEST_SUMMARY_COLUMNS:
        row1.append(data.get(col, ""))
    ws1.append(row1)
    
    # Apply verdict color to the last row
    last_row = ws1.max_row
    verdict_col = TEST_SUMMARY_COLUMNS.index("overall_verdict") + 1
    verdict_value = data.get("overall_verdict", "")
    if verdict_value in VERDICT_COLORS:
        ws1.cell(row=last_row, column=verdict_col).fill = VERDICT_COLORS[verdict_value]
    
    # Add to Sheet 2: Problem Testcases
    ws2 = wb[SHEET_PROBLEM_TESTCASES]
    problem_testcases = data.get("problem_testcases", {})
    total_solved = (
        problem_testcases.get("easy_solved", 0) +
        problem_testcases.get("medium_solved", 0) +
        problem_testcases.get("hard_solved", 0)
    )
    total_problems = (
        problem_testcases.get("easy_total", 2) +
        problem_testcases.get("medium_total", 9) +
        problem_testcases.get("hard_total", 9)
    )
    
    row2 = [
        data.get("candidate_id", ""),
        data.get("name", ""),
        problem_testcases.get("easy_solved", 0),
        problem_testcases.get("easy_total", 2),
        problem_testcases.get("medium_solved", 0),
        problem_testcases.get("medium_total", 9),
        problem_testcases.get("hard_solved", 0),
        problem_testcases.get("hard_total", 9),
        total_solved,
        total_problems
    ]
    ws2.append(row2)
    
    # Add to Sheet 3: Testcase Details
    ws3 = wb[SHEET_TESTCASE_DETAILS]
    problem_scores = data.get("problem_scores", {})
    row3 = [
        data.get("candidate_id", ""),
        data.get("name", ""),
        problem_scores.get("P1_py", 0),
        problem_scores.get("P2_py", 0),
        problem_scores.get("P3_py", 0),
        problem_scores.get("P4_py", 0),
        problem_scores.get("P5_py", 0),
        problem_scores.get("P6_sql", 0),
        problem_scores.get("P7_sql", 0),
        problem_scores.get("P8_sql", 0),
        problem_scores.get("P9_sql", 0),
        problem_scores.get("P10_sql", 0)
    ]
    ws3.append(row3)
    
    wb.save(EXCEL_FILE_PATH)
    wb.close()
    
    return {"success": True, "candidate_id": data.get("candidate_id")}


def export_excel(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    verdict: Optional[str] = None,
    submission_type: Optional[str] = None
) -> bytes:
    """Export filtered results as Excel file with formatting"""
    results = read_all_results(date_from, date_to, verdict, submission_type)
    
    wb = Workbook()
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def style_sheet(ws: Worksheet, headers: List[str]):
        """Apply consistent styling to a sheet"""
        # Style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Freeze top row
        ws.freeze_panes = "A2"
        
        # Auto-filter
        ws.auto_filter.ref = ws.dimensions
    
    def auto_fit_columns(ws: Worksheet):
        """Auto-fit column widths"""
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    # Sheet 1: Test Summary
    ws1 = wb.active
    ws1.title = SHEET_TEST_SUMMARY
    style_sheet(ws1, TEST_SUMMARY_COLUMNS)
    
    for row_idx, record in enumerate(results, 2):
        for col_idx, col_name in enumerate(TEST_SUMMARY_COLUMNS, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            
            # Apply verdict color
            if col_name == "overall_verdict":
                verdict_val = record.get(col_name, "")
                if verdict_val in VERDICT_COLORS:
                    cell.fill = VERDICT_COLORS[verdict_val]
    
    auto_fit_columns(ws1)
    
    # Sheet 2: Problem Testcases
    ws2 = wb.create_sheet(SHEET_PROBLEM_TESTCASES)
    style_sheet(ws2, PROBLEM_TESTCASES_COLUMNS)
    
    for row_idx, record in enumerate(results, 2):
        problem_testcases = record.get("problem_testcases", {})
        total_solved = (
            problem_testcases.get("easy_solved", 0) +
            problem_testcases.get("medium_solved", 0) +
            problem_testcases.get("hard_solved", 0)
        )
        total_problems = (
            problem_testcases.get("easy_total", 2) +
            problem_testcases.get("medium_total", 9) +
            problem_testcases.get("hard_total", 9)
        )
        
        row_data = [
            record.get("candidate_id", ""),
            record.get("name", ""),
            problem_testcases.get("easy_solved", 0),
            problem_testcases.get("easy_total", 2),
            problem_testcases.get("medium_solved", 0),
            problem_testcases.get("medium_total", 9),
            problem_testcases.get("hard_solved", 0),
            problem_testcases.get("hard_total", 9),
            total_solved,
            total_problems
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
    
    auto_fit_columns(ws2)
    
    # Sheet 3: Testcase Details
    ws3 = wb.create_sheet(SHEET_TESTCASE_DETAILS)
    style_sheet(ws3, TESTCASE_DETAILS_COLUMNS)
    
    for row_idx, record in enumerate(results, 2):
        problem_scores = record.get("problem_scores", {})
        row_data = [
            record.get("candidate_id", ""),
            record.get("name", ""),
            problem_scores.get("P1_py", 0),
            problem_scores.get("P2_py", 0),
            problem_scores.get("P3_py", 0),
            problem_scores.get("P4_py", 0),
            problem_scores.get("P5_py", 0),
            problem_scores.get("P6_sql", 0),
            problem_scores.get("P7_sql", 0),
            problem_scores.get("P8_sql", 0),
            problem_scores.get("P9_sql", 0),
            problem_scores.get("P10_sql", 0)
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
    
    auto_fit_columns(ws3)
    
    # Save to bytes
    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_sample_data():
    """Create sample data for 5 candidates"""
    sample_candidates = [
        {
            "candidate_id": "C001",
            "name": "Rahul Kumar",
            "email": "rahul@test.com",
            "phone": "9876543210",
            "test_date": "2026-02-20",
            "login_time": "09:00",
            "submit_time": "10:45",
            "submission_type": "Manual",
            "time_taken_min": 105,
            "total_questions": 10,
            "python_questions": 5,
            "sql_questions": 5,
            "python_score": 85,
            "sql_score": 78,
            "problem_testcases": {
                "easy_solved": 2, "easy_total": 2,
                "medium_solved": 7, "medium_total": 9,
                "hard_solved": 5, "hard_total": 9
            },
            "problem_scores": {
                "P1_py": 5, "P2_py": 5, "P3_py": 4, "P4_py": 4, "P5_py": 5,
                "P6_sql": 5, "P7_sql": 4, "P8_sql": 4, "P9_sql": 3, "P10_sql": 4
            }
        },
        {
            "candidate_id": "C002",
            "name": "Priya Sharma",
            "email": "priya@test.com",
            "phone": "9876543211",
            "test_date": "2026-02-20",
            "login_time": "09:00",
            "submit_time": "11:00",
            "submission_type": "Auto",
            "time_taken_min": 120,
            "total_questions": 10,
            "python_questions": 5,
            "sql_questions": 5,
            "python_score": 45,
            "sql_score": 52,
            "problem_testcases": {
                "easy_solved": 2, "easy_total": 2,
                "medium_solved": 4, "medium_total": 9,
                "hard_solved": 2, "hard_total": 9
            },
            "problem_scores": {
                "P1_py": 5, "P2_py": 3, "P3_py": 2, "P4_py": 2, "P5_py": 3,
                "P6_sql": 4, "P7_sql": 3, "P8_sql": 3, "P9_sql": 2, "P10_sql": 2
            }
        },
        {
            "candidate_id": "C003",
            "name": "Amit Patel",
            "email": "amit@test.com",
            "phone": "9876543212",
            "test_date": "2026-02-21",
            "login_time": "10:00",
            "submit_time": "11:30",
            "submission_type": "Manual",
            "time_taken_min": 90,
            "total_questions": 10,
            "python_questions": 5,
            "sql_questions": 5,
            "python_score": 32,
            "sql_score": 28,
            "problem_testcases": {
                "easy_solved": 1, "easy_total": 2,
                "medium_solved": 3, "medium_total": 9,
                "hard_solved": 1, "hard_total": 9
            },
            "problem_scores": {
                "P1_py": 3, "P2_py": 2, "P3_py": 2, "P4_py": 1, "P5_py": 2,
                "P6_sql": 3, "P7_sql": 2, "P8_sql": 1, "P9_sql": 1, "P10_sql": 1
            }
        },
        {
            "candidate_id": "C004",
            "name": "Sneha Reddy",
            "email": "sneha@test.com",
            "phone": "9876543213",
            "test_date": "2026-02-21",
            "login_time": "14:00",
            "submit_time": "15:50",
            "submission_type": "Manual",
            "time_taken_min": 110,
            "total_questions": 10,
            "python_questions": 5,
            "sql_questions": 5,
            "python_score": 92,
            "sql_score": 88,
            "problem_testcases": {
                "easy_solved": 2, "easy_total": 2,
                "medium_solved": 8, "medium_total": 9,
                "hard_solved": 7, "hard_total": 9
            },
            "problem_scores": {
                "P1_py": 5, "P2_py": 5, "P3_py": 5, "P4_py": 4, "P5_py": 5,
                "P6_sql": 5, "P7_sql": 5, "P8_sql": 4, "P9_sql": 5, "P10_sql": 4
            }
        },
        {
            "candidate_id": "C005",
            "name": "Karthik Menon",
            "email": "karthik@test.com",
            "phone": "9876543214",
            "test_date": "2026-02-22",
            "login_time": "09:30",
            "submit_time": "11:30",
            "submission_type": "Auto",
            "time_taken_min": 120,
            "total_questions": 10,
            "python_questions": 5,
            "sql_questions": 5,
            "python_score": 58,
            "sql_score": 62,
            "problem_testcases": {
                "easy_solved": 2, "easy_total": 2,
                "medium_solved": 5, "medium_total": 9,
                "hard_solved": 3, "hard_total": 9
            },
            "problem_scores": {
                "P1_py": 5, "P2_py": 4, "P3_py": 3, "P4_py": 3, "P5_py": 3,
                "P6_sql": 4, "P7_sql": 4, "P8_sql": 3, "P9_sql": 3, "P10_sql": 3
            }
        }
    ]
    
    for candidate in sample_candidates:
        add_result(candidate)
    
    return len(sample_candidates)
